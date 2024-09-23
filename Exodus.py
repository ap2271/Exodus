# Third-Party Libraries and Licenses
#
# This application uses the following third-party libraries:
#
# 1. pandas 2.2.2
#    - License: BSD-3-Clause
#    - Copyright (c) 2008-2024, the pandas development team
#
# 2. pythonping 1.1.4
#    - License: MIT License
#    - Copyright (c) 2021 pythonping
#
# 3. openpyxl   3.1.2
#    - License: MIT License
#    - Copyright (c) 2010-2024, openpyxl developers
#
# 4. PyQt5  5.15.10, QT5 5.15.2
#    - License: GPL v3 or commercial license
#    - Copyright (c) 2001-2024, Riverbank Computing Limited
#
# Full license texts are included in the LICENSES directory of this application.

import os
import sys
import threading
import logging
import re
import shutil
import configparser
import getpass
import subprocess
import pandas as pd
from datetime import datetime
from ftplib import FTP
from pathlib import Path
from pythonping import ping
import openpyxl
import traceback
from PyQt5.QtCore import QObject, pyqtSignal, pyqtSlot, QMutex, Qt, QTimer, QUrl, QSize
from PyQt5.QtGui import (
    QIcon,
    QPixmap,
    QDesktopServices,
    QBrush,
    QColor,
    QImage,
    QPalette,
)
from PyQt5.QtWidgets import (
    QApplication,
    QWidget,
    QHBoxLayout,
    QVBoxLayout,
    QPushButton,
    QFileDialog,
    QLabel,
    QListWidget,
    QGroupBox,
    QAction,
    QSplashScreen,
    QCheckBox,
    QGridLayout,
    QMessageBox,
    QMainWindow,
    QDialog,
    QTextBrowser,
    QComboBox,
    QScrollArea,
    QTabWidget,
    QTableWidget,
    QLineEdit,
    QTextEdit,
    QTableWidgetItem,
    QStyledItemDelegate,
    QSpacerItem,
    QSizePolicy,
    QStyle,
    QStyleOptionButton,
    QListWidgetItem,
    QProgressDialog,
    QToolButton,
    QAbstractItemView,
    QProgressBar,
    QStatusBar,
)

TotalRobots = 0
ThreadCount = 0


def resource_path(relative_path):
    if hasattr(sys, "_MEIPASS"):  # PyInstaller
        base_path = sys._MEIPASS
    elif hasattr(sys, "_MEIPASS2"):  # Nuitka
        base_path = sys._MEIPASS2
    else:
        base_path = os.path.abspath(".")

    full_path = os.path.join(base_path, relative_path)

    # Debug: Print the full path to check if it resolves correctly
    print(f"Looking for resource at: {full_path}")

    # Check if the file exists
    if not os.path.exists(full_path):
        print(f"Resource not found: {full_path}")

    return full_path


class FTPBackup(QMainWindow):  # Changed from QWidget to QMainWindow
    def __init__(self):
        super().__init__()
        self.initUI()

    def initUI(self):

        # Main Window
        try:
            self.logger = None  # Initialize logger to None
            global ThreadCount
            # Check for key combination (Ctrl + L) to turn on logger
            toggle_logger_sequence = "F9"
            toggle_logger_action = QAction(self)
            toggle_logger_action.setShortcut(toggle_logger_sequence)
            toggle_logger_action.triggered.connect(self.toggle_logger)
            self.addAction(toggle_logger_action)
            self.logger_action = toggle_logger_action
            self.setWindowIcon(QIcon(resource_path("MAGNAlogo.png")))
            self.setWindowTitle("Exodus")
            self.setFocusPolicy(Qt.StrongFocus)
            self.setMinimumSize(700, 650)
            self.setMaximumWidth(700)

            self.status_bar = QStatusBar()
            self.setStatusBar(self.status_bar)
            # Create a label for the icon
            self.icon_label = QLabel()
            self.icon_label.setPixmap(
                QIcon(resource_path("MAGNA.png")).pixmap(100, 100)
            )

            # Add the label to the status bar as a permanent widget
            self.status_bar.addPermanentWidget(self.icon_label)

            # oImage = QImage("MAGNA.ico")
            # sImage = oImage.scaled(QSize(700, 350))  # resize Image to widgets size
            # palette = QPalette()
            # palette.setBrush(QPalette.Window, QBrush(sImage))
            # self.setPalette(palette)

            # Tab Widget
            self.tab_widget = QTabWidget()
            self.setCentralWidget(self.tab_widget)

            self.Project = QWidget()
            self.Status = QWidget()

            self.tab_widget.addTab(self.Project, "Main")
            self.tab_widget.addTab(self.Status, "Status")

            self.main_layout = QHBoxLayout(self.Project)

            # Create an opacity effect
            # opacity_effect = QGraphicsOpacityEffect()
            # opacity_effect.setOpacity(
            #     0.5
            # )  # Value between 0 (fully transparent) and 1 (fully opaque)
            #
            # # Apply the effect to the widget
            # self.tab_widget.setGraphicsEffect(opacity_effect)

            self.mainleft_layout = QVBoxLayout()

            self.info_box = QTableWidget()
            self.info_box.setStyleSheet("QTableWidget { border-radius: 10px; }")

            self.info_box.setColumnCount(3)
            self.info_box.setColumnWidth(0, 100)
            self.info_box.setColumnWidth(1, 150)
            self.info_box.verticalHeader().setVisible(False)
            self.info_box.setGridStyle(False)
            self.info_box.setSelectionMode(QTableWidget.NoSelection)
            self.info_box.setEditTriggers(QTableWidget.NoEditTriggers)
            self.info_box.setFocusPolicy(Qt.NoFocus)
            self.info_box.setHorizontalHeaderLabels(
                ["Robot Name", "IP Address", "Status"]
            )

            self.mainleft_layout.addWidget(self.info_box)

            self.mainright_layout = QVBoxLayout()
            selection_group_box = QGroupBox("Simultaneous Backups")
            selection_group_box.setToolTip(
                "Number of Robots that can be set for backup at the same time\n Higher number will speed up the backup process."
            )
            selection_group_box.setFixedSize(200, 60)
            selection_group_boxlayout = QVBoxLayout()

            # Completed Group Box
            # Left Layout
            # Thread Count dropdown menu

            # Drop-down menu for selecting simultaneous thread counts
            self.thread_count_combobox = QComboBox()
            self.thread_count_combobox.setEditable(False)
            self.thread_count_combobox.setFixedSize(100, 20)
            self.thread_count_combobox = QComboBox()
            thread_counts = [1, 5, 10, 20, 30, 50, 100, 150, 180]
            self.thread_count_combobox.addItems([str(count) for count in thread_counts])
            self.thread_count_combobox.setCurrentIndex(2)  # Default to 10 threads
            self.thread_count_combobox.currentIndexChanged.connect(
                self.update_thread_count
            )

            selection_group_boxlayout.addWidget(
                self.thread_count_combobox
            )  # Add the combo box to the left layout
            selection_group_box.setLayout(selection_group_boxlayout)
            self.extension_checkboxes_group_box = QGroupBox(
                "Select Extensions to Download (/MD:)"
            )
            self.extension_checkboxes_group_box.setToolTip(
                "/MD: Backup\n"
                "Selection of file type to be backup.\n"
                "All Files Option will backup all the files in\n"
                "directory regardless of it's extension type.\n"
                "individual selection will only download that\n"
                "particular file type."
            )
            self.extension_checkboxes_group_box.setFixedSize(200, 250)
            self.extension_checkboxes_layout = (
                QGridLayout()
            )  # Changed QVBoxLayout to QGridLayout
            self.extension_checkboxes = []
            extensions = [
                ".zip",
                ".stm",
                ".tx",
                ".pc",
                ".vda",
                ".dg",
                ".dt",
                ".io",
                ".sv",
                ".va",
                ".vr",
                ".ls",
                ".tp",
            ]
            self.select_all_checkbox = QCheckBox("*.* All Files (MD:)", self)
            self.extension_checkboxes_layout.addWidget(
                self.select_all_checkbox, 0, 0
            )  # Add checkbox at position (0, 0)
            row = 1  # Start from the second row
            column = 0  # Start from the first column
            for ext in extensions:
                if row < 7:
                    checkbox = QCheckBox(ext, self)
                    self.extension_checkboxes_layout.addWidget(
                        checkbox, row, column
                    )  # Add checkbox at specified row and column
                    checkbox.stateChanged.connect(self.on_extension_checkbox_toggled)
                    self.select_all_checkbox.stateChanged.connect(
                        self.toggle_all_if_select_all
                    )
                    self.extension_checkboxes.append(checkbox)
                    row += 1  # Move to the next row for the next checkbox
                else:
                    row = 0
                    column += 1
                    checkbox = QCheckBox(ext, self)
                    self.extension_checkboxes_layout.addWidget(
                        checkbox, row, column
                    )  # Add checkbox at specified row and column
                    checkbox.stateChanged.connect(self.on_extension_checkbox_toggled)
                    self.select_all_checkbox.stateChanged.connect(
                        self.toggle_all_if_select_all
                    )
                    self.extension_checkboxes.append(checkbox)
                    row += 1  # Move to the next row for the next checkbox
            # self.select_all_checkbox.stateChanged.connect(self.toggle_select_all)  removed  in V3.5

            self.select_all_checkbox.setChecked(True)
            self.extension_checkboxes_group_box.setLayout(
                self.extension_checkboxes_layout
            )
            backupButton_group_box = QGroupBox()
            backupButton_group_box.setFixedSize(200, 40)
            backupButton_group_boxlayout = QVBoxLayout()
            self.backup_button = QPushButton("Initiate Backups")
            self.backup_button.setStatusTip("Open Project to start backups")

            self.backup_button.setEnabled(False)
            self.backup_button.setFixedSize(150, 25)
            self.backup_button.clicked.connect(self.start_backup)
            backupButton_group_boxlayout.addWidget(
                self.backup_button,
                alignment=Qt.AlignCenter,
            )
            backupButton_group_box.setLayout(backupButton_group_boxlayout)

            statuscheck_group_box = QGroupBox()
            statuscheck_group_box.setFixedSize(200, 100)
            statuscheck_group_boxlayout = QVBoxLayout()
            self.statuscheck_button = QPushButton("Check Status")
            self.statuscheck_button.setToolTip(
                "Ping all the robots in the project and show the current status."
            )
            self.statuscheck_button.clicked.connect(self.check_robot_status_button)
            self.statuscheck_button.setStatusTip("Ping Robots to check status")
            self.statuscheck_button.setFixedSize(150, 25)
            self.statuscheck_button.setEnabled(False)
            self.selectall_checkbox = QCheckBox("Select All", self)
            self.selectall_checkbox.setToolTip(
                "Select all the robots in the project for the backup\n"
                "Regardless of it's status."
            )
            self.selectall_checkbox.clicked.connect(self.select_all_robots)
            self.selectonline_checkbox = QCheckBox("Select Online", self)
            self.selectonline_checkbox.setToolTip(
                "Select only online robots for the Backup."
            )
            self.selectonline_checkbox.clicked.connect(self.select_robots_by_status)
            statuscheck_group_boxlayout.addWidget(
                self.selectall_checkbox,
            )
            statuscheck_group_boxlayout.addWidget(
                self.selectonline_checkbox,
            )
            statuscheck_group_boxlayout.addWidget(
                self.statuscheck_button,
                alignment=Qt.AlignCenter,
            )
            statuscheck_group_box.setLayout(statuscheck_group_boxlayout)

            self.mainright_layout.addWidget(statuscheck_group_box)
            self.mainright_layout.addWidget(selection_group_box)
            self.mainright_layout.addWidget(self.extension_checkboxes_group_box)
            self.mainright_layout.addWidget(backupButton_group_box)

            self.main_layout.addLayout(self.mainleft_layout)
            self.main_layout.addLayout(self.mainright_layout)

            self.project_layout = QHBoxLayout(
                self.Status
            )  # Set layout on central widget
            self.left_layout = QVBoxLayout()

            # Completed Group Box
            self.completed_group_box = QGroupBox("Completed Backups")
            self.completed_layout = QVBoxLayout()
            self.completed_group_box.setFixedWidth(300)
            self.completed_list_widget = QListWidget()
            self.completed_layout.addWidget(self.completed_list_widget)
            self.completed_group_box.setLayout(self.completed_layout)
            self.left_layout.addWidget(self.completed_group_box)
            self.project_layout.addLayout(self.left_layout)
            self.right_layout = QVBoxLayout()
            self.scheduled_group_box = QGroupBox("Scheduled and Running Backups")
            self.scheduled_layout = QVBoxLayout()
            self.scheduled_group_box.setFixedWidth(385)
            self.scheduled_list_widget = QListWidget()
            self.scheduled_layout.addWidget(self.scheduled_list_widget)
            self.scheduled_group_box.setLayout(self.scheduled_layout)
            self.right_layout.addWidget(self.scheduled_group_box)

            # Button layout
            button_layout = QHBoxLayout()
            self.textbox = QLabel(self)
            self.textbox.setText("File: ")
            self.textbox.setFixedSize(20, 25)
            self.Filename = QLabel(self)
            self.Filename.setFixedSize(200, 25)
            self.open_folder_button = QPushButton("Open Backup")
            self.open_folder_button.setFixedSize(80, 25)
            self.open_folder_button.setVisible(False)  # Initially disabled
            self.open_folder_button.clicked.connect(self.open_backup_folder)
            button_layout.addWidget(self.textbox)
            button_layout.addWidget(self.Filename)
            button_layout.addWidget(self.open_folder_button)  # Add the new button
            self.right_layout.addLayout(button_layout)
            self.project_layout.addLayout(self.right_layout)

            self.worker = Worker()
            self.worker.progress_signal.connect(self.update_progress)
            self.create_menu_bar()  # Create menu bar
        except Exception as e:
            if self.logger:
                self.logger.error(f"Error initializing UI: {str(e)}")
            else:
                print(f"Error initializing UI: {str(e)}")

    def on_extension_checkbox_toggled(self, state):
        if state == Qt.Checked:
            # Uncheck the "All Files" checkbox if it is currently checked
            if self.select_all_checkbox.isChecked():
                self.select_all_checkbox.setChecked(False)

    def toggle_all_if_select_all(self, state):

        for checkbox in self.extension_checkboxes:
            checkbox.setChecked(False)

    def update_thread_count(self):

        global Thread_count
        selected_index = self.thread_count_combobox.currentIndex()
        thread_counts = [1, 5, 10, 20, 30, 50, 100, 150, 180]
        Thread_count = thread_counts[selected_index]
        return Thread_count

    def toggle_logger(self):

        if self.logger is None:
            exe_dir = os.path.dirname(sys.executable)
            log_file_path = os.path.join(exe_dir, "crash_log_UI.dmp")
            try:
                logging.basicConfig(
                    filename=log_file_path,
                    level=logging.DEBUG,
                    format="%(asctime)s - %(levelname)s - %(message)s",
                )
                self.logger = logging.getLogger(__name__)
                self.logger.info("Logger turned on")
                self.update_scheduled_list("logger_on")
            except Exception as e:
                print("Error occurred while setting up logger:", e)
        else:
            logging.shutdown()
            self.logger = None
            print("Logger turned off")

    def open_backup_folder(self):

        try:
            if self.logger:
                self.logger.info("Opening backup folder.")
            else:
                print("Logger is not initialized.")
            if hasattr(self, "main_folder") and os.path.exists(self.main_folder):
                try:
                    if self.logger:
                        self.logger.info("Opening backup folder.")

                    os.startfile(self.main_folder)  # For Windows
                except AttributeError:
                    if self.logger:
                        self.logger.error(
                            "Error: os.startfile is not supported on this platform."
                        )

                    print("Error: os.startfile is not supported on this platform.")
            else:
                QMessageBox.warning(
                    self, "Folder Not Found", "Backup folder not found."
                )
        except Exception as e:
            if self.logger:
                self.logger.error(f"Error opening backup folder: {str(e)}")

    def create_menu_bar(self):

        menubar = self.menuBar()

        fileMenu = menubar.addMenu("File")
        helpMenu = menubar.addMenu("Help")

        exitAction = QAction("&Exit", self)
        exitAction.setShortcut("Ctrl+Q")
        exitAction.setStatusTip("Exit application")
        exitAction.triggered.connect(self.close)

        createprojectAction = QAction("&New", self)
        createprojectAction.setShortcut("Ctrl+N")
        createprojectAction.setStatusTip("New Project")
        createprojectAction.triggered.connect(self.showNewFileDialog)

        editprojectAction = QAction("&Edit", self)
        editprojectAction.setShortcut("Ctrl+E")
        editprojectAction.setStatusTip("Edit Project")
        editprojectAction.triggered.connect(self.showeditFileDialog)

        OpenprojectAction = QAction("&Open", self)
        OpenprojectAction.setShortcut("Ctrl+O")
        OpenprojectAction.setStatusTip("Open Project")
        OpenprojectAction.triggered.connect(self.show_open_project_dialog)

        aboutAction = QAction("&About", self)
        aboutAction.setStatusTip("About this application")
        aboutAction.triggered.connect(self.show_about_dialog)

        guideAction = QAction("&Guide", self)
        guideAction.setStatusTip("User Guide")
        guideAction.triggered.connect(self.show_guide_dialog)

        fileMenu.addAction(createprojectAction)
        fileMenu.addAction(editprojectAction)
        fileMenu.addAction(OpenprojectAction)
        fileMenu.addAction(exitAction)
        helpMenu.addAction(aboutAction)
        helpMenu.addAction(guideAction)

    def showNewFileDialog(self):
        self.new_dialog = ProjectConfigEditor()
        self.new_dialog.exec_()

    def showeditFileDialog(self):
        self.edit_dialog = ProjectConfigEditor()
        self.edit_dialog.editProject()
        self.edit_dialog.exec_()

    def show_open_project_dialog(self):
        user_documents_path = os.path.join("C:\\Users", getpass.getuser(), "Documents")
        robot_backup_path = os.path.join(user_documents_path, "Exodus")

        if os.path.exists(robot_backup_path):
            openfile = QFileDialog.getOpenFileName(
                self, "Open Project", robot_backup_path, "Project Files (*.ini)"
            )
            if openfile[0]:
                self.open_project(openfile[0])
        else:
            QMessageBox.warning(self, "Error", "create blank project first")

    def open_project(self, project_file_path):

        if not os.path.exists(project_file_path):
            QMessageBox.warning(self, "Error", "Project file does not exist.")
            return

        config = configparser.ConfigParser()
        try:
            config.read(project_file_path)
        except Exception as e:
            QMessageBox.warning(self, "Error", f"Failed to read project file: {e}")
            return

        backup_directory = config.get("General", "BackupDirectory", fallback="")
        tempdir = os.path.splitext(os.path.basename(project_file_path))[0]
        # print(backup_directory + "\\" + tempdir)
        backup_directory = backup_directory + "\\" + tempdir

        robots_section = config.items("Robots")
        # print(robots_section)

        self.populate_data(backup_directory, robots_section, project_file_path)

        if not robots_section == []:
            self.backup_button.setEnabled(True)

    from PyQt5.QtWidgets import QProgressDialog

    def populate_data(self, backup_directory, robots_section, project_file_path):
        try:
            # Initialize QProgressDialog
            progress_dialog = QProgressDialog(
                "Loading data...", None, 0, len(robots_section), self
            )
            progress_dialog.setWindowModality(Qt.WindowModal)
            progress_dialog.setWindowTitle("Importing")
            progress_dialog.setMinimumDuration(0)
            progress_dialog.setValue(0)
            progress_dialog.setCancelButton(None)
            progress_dialog.setWindowFlags(
                progress_dialog.windowFlags() & ~Qt.WindowCloseButtonHint
            )  # Disable the close button
            self.info_box.clearContents()  # Clear existing contents
            self.info_box.setRowCount(len(robots_section))

            row = 0
            for robot_name, ip_address in robots_section:
                if progress_dialog.wasCanceled():
                    break

                ip_item = QTableWidgetItem(ip_address)
                ip_item.setTextAlignment(Qt.AlignCenter)  # Align the item in the center

                # Create a custom widget to hold the checkbox and text
                widget = QWidget()
                layout = QHBoxLayout()
                checkbox = QCheckBox()
                label = QLabel(robot_name.upper())

                layout.addWidget(checkbox)
                spacer_item = QSpacerItem(2, 2, QSizePolicy.Fixed, QSizePolicy.Minimum)
                layout.addItem(spacer_item)
                layout.addWidget(label)
                layout.setContentsMargins(0, 0, 0, 0)
                widget.setLayout(layout)

                # Set the custom widget as the cell widget in the first column
                self.info_box.setCellWidget(row, 0, widget)

                # Set the IP address as a regular item in the second column
                self.info_box.setItem(row, 1, ip_item)

                self.info_box.setAlternatingRowColors(True)

                is_online = self.is_robot_alive(ip_item.text())
                status_label = QLabel()
                status_label.setText("Online" if is_online else "Offline")
                status_label.setStyleSheet(
                    f"background-color: {'#7ABA78' if is_online else '#F2613F'}; color: white; border-radius: 5px; padding: 2px 5px;"
                )
                self.info_box.setCellWidget(row, 2, status_label)

                row += 1
                progress_dialog.setValue(row)  # Update the progress dialog

            self.backup_path = backup_directory
            self.Filename.setText(os.path.basename(project_file_path))
            self.Filename.setStyleSheet(
                "background-color: #005FB8; color: white; border-radius: 5px; font-weight: bold; padding: 5px;"
            )
            self.setWindowTitle(
                "Exodus         "
                + "Current Project:  "
                + os.path.splitext(os.path.basename(project_file_path))[0]
            )

            progress_dialog.setValue(len(robots_section))  # Close the progress dialog
            self.statuscheck_button.setEnabled(True)
        except Exception as e:
            if self.logger:
                self.logger.error(f"Error populating data: {str(e)}")
            else:
                print(f"Error populating data: {str(e)}")
            progress_dialog.cancel()

    def select_robots_by_status(self):
        # Get the status of checkboxes
        select_online_checked = self.selectonline_checkbox.isChecked()

        # Iterate through rows in the main table
        for row in range(self.info_box.rowCount()):
            robot_name_widget = self.info_box.cellWidget(row, 0)
            robot_name_checkbox = robot_name_widget.findChild(QCheckBox)

            status_label = self.info_box.cellWidget(row, 2)
            status_text = status_label.text()

            # Check if the robot should be selected based on checkbox states and status
            if (
                select_online_checked
                and status_text == "Online"
                and robot_name_checkbox
            ):
                robot_name_checkbox.setChecked(True)
            else:
                robot_name_checkbox.setChecked(False)

    def select_all_robots(self, state):
        # Get the state of the "Select All" checkbox
        select_all_checked = self.selectall_checkbox.isChecked()

        # Iterate through rows in the main table
        for row in range(self.info_box.rowCount()):
            # Find the checkbox in the first column of each row
            checkbox_item = self.info_box.cellWidget(row, 0)
            if isinstance(checkbox_item, QWidget):
                checkbox = checkbox_item.findChild(QCheckBox)
                # Check or uncheck the checkbox based on the state of the "Select All" checkbox
                if checkbox:
                    checkbox.setChecked(select_all_checked)

    def check_robot_status_button(self):
        # Initialize QProgressDialog
        progress_dialog = QProgressDialog(
            "Checking robot status...", None, 0, self.info_box.rowCount(), self
        )
        progress_dialog.setWindowModality(Qt.WindowModal)
        progress_dialog.setWindowTitle("Querying Robots")
        progress_dialog.setMinimumDuration(0)
        progress_dialog.setCancelButton(None)
        progress_dialog.setWindowFlags(
            progress_dialog.windowFlags() & ~Qt.WindowCloseButtonHint
        )  # Disable the close button

        # Iterate through rows in the main table
        for row in range(self.info_box.rowCount()):
            if progress_dialog.wasCanceled():
                break

            robot_name_widget = self.info_box.cellWidget(row, 0)
            robot_name_checkbox = (
                robot_name_widget.layout().itemAt(0).widget()
            )  # Find the checkbox
            ip_address_item = self.info_box.item(row, 1)

            if robot_name_checkbox and ip_address_item:
                ip_address = ip_address_item.text()
                is_online = self.is_robot_alive(ip_address)
                status_label = QLabel()
                status_label.setText("Online" if is_online else "Offline")
                status_label.setStyleSheet(
                    f"background-color: {'#7ABA78' if is_online else '#F2613F'}; color: white; border-radius: 5px; padding: 2px 5px;"
                )
                self.info_box.setCellWidget(row, 2, status_label)

            self.info_box.setCellWidget(row, 2, status_label)

            progress_dialog.setValue(row + 1)  # Update the progress dialog

        progress_dialog.setValue(self.info_box.rowCount())  # Close the progress dialog

    def compile_robot_info(self):
        try:
            self.robot_info = []
            for row in range(self.info_box.rowCount()):
                robot_name_widget = self.info_box.cellWidget(row, 0)
                robot_name_checkbox = robot_name_widget.findChild(QCheckBox)
                if robot_name_checkbox and robot_name_checkbox.isChecked():
                    robot_name_label = robot_name_widget.layout().itemAt(2).widget()
                    robot_name = (
                        robot_name_label.text().strip()
                        if isinstance(robot_name_label, QLabel)
                        else ""
                    )
                    ip_item = self.info_box.item(row, 1)
                    ip_address = ip_item.text()

                    # Check if the robot is selected
                    if robot_name and ip_address:
                        self.robot_info.append((robot_name, ip_address))

            # Print or log the compiled robot_info for verification
            # print("Compiled robot_info:", self.robot_info)

        except Exception as e:
            if self.logger:
                self.logger.error(f"Error compiling robot info: {str(e)}")
            else:
                print(f"Error compiling robot info: {str(e)}")

    def show_about_dialog(self):
        image_path1 = resource_path("Robot.ico")
        image_path2 = resource_path("MAGNA.png")  # Path to the new image
        about_text = (
            "<html><body>"
            "<div style='display:flex; align-items: flex-start;'>"
            "<div>"
            "<img src='{}' alt='Logo2' style='width: 50px; height: 50px; margin-left: 10px;'>"
            "</div>"
            "<div style='margin-left: 10px;'>"
            "<h2 style='color:black;'>Exodus Robot Backup Tool</h2>"
            "<p>Author: MSVNBAPATEL</p>"
            "<p>Email: <a href='mailto:amit.patel1@magna.com'>amit.patel1@magna.com</a></p>"
            "<p>Runtime version: 3.7</p>"
            "</div>"
            "</div>"
            "</body></html>"
        ).format(
            image_path2
        )  # Format the HTML to include both image paths

        about_box = QMessageBox()
        about_box.setWindowTitle("About Exodus")
        about_box.setText(about_text)
        about_box.setFixedSize(800, 800)

        # Add a "See Change Log" button
        see_changelog_button = about_box.addButton(
            "View Change Log", QMessageBox.ActionRole
        )

        # Connect the button to a method to open the change log
        see_changelog_button.clicked.connect(self.show_changelog)

        # Set text format to allow HTML rendering
        about_box.setTextFormat(Qt.RichText)
        about_box.setTextInteractionFlags(Qt.TextBrowserInteraction)

        about_box.exec_()

    def show_changelog(self):

        # Load the content of the change log text file
        changelog_text = ""
        try:
            if self.logger:
                self.logger.info("Reading ChangeLog.txt")
            # with open("ChangeLog.txt", "r") as file:
            with open(resource_path("ChangeLog.txt"), "r") as file:
                changelog_text = file.read()
        except Exception as e:
            if self.logger:
                self.logger.error(f"Error reading ChangeLog.txt: {str(e)}")

        # Create a new dialog to display the change log content
        changelog_dialog = QDialog(self)
        changelog_dialog.setWindowTitle("Change Log")
        changelog_dialog.setFixedSize(600, 400)

        # Use QTextBrowser to display the text content
        text_browser = QTextBrowser()
        text_browser.setPlainText(changelog_text)

        # Create a scroll area and set the QTextBrowser as its widget
        scroll_area = QScrollArea()
        scroll_area.setWidgetResizable(True)
        scroll_area.setWidget(text_browser)

        # Layout for the dialog
        layout = QVBoxLayout()
        layout.addWidget(scroll_area)
        changelog_dialog.setLayout(layout)

        changelog_dialog.exec_()

    def show_guide_dialog(self):

        pdf_path = resource_path("Manual.pdf")

        # Open the PDF file
        QDesktopServices.openUrl(QUrl.fromLocalFile(pdf_path))

    def closeEvent(self, event):

        # self.terminate_threads()
        event.accept()  # Accept the close event

    def select_file(self):

        self.scheduled_list_widget.clear()
        file_dialog = QFileDialog(self)
        file_path, _ = file_dialog.getOpenFileName(
            self, "Select File", "", "Text files (*.txt)"
        )
        if file_path:
            self.file_path = file_path
            self.Filename.setText(os.path.basename(file_path))
            self.Filename.setStyleSheet(
                "background-color: #005FB8 ;color: white;"
                " border-radius: 5px ;font-weight: bold;"
                "padding: 5px 5px 5px 5px;"
            )
            try:
                if self.logger:
                    self.logger.info("Reading file")

                self.read_robot_info(file_path)
                self.check_robot_status()
                if self.robot_info == []:
                    self.backup_button.setEnabled(False)
                else:
                    self.backup_button.setEnabled(True)
            except Exception as e:
                if self.logger:
                    self.logger.error(f"Error reading file: {str(e)}")

    def check_robot_status(self):

        for robot_name, ip_address in self.robot_info:
            if ip_address:
                if not self.is_robot_alive(ip_address):

                    self.remove_robot_info(robot_name)
                    self.update_scheduled_list(
                        f"{robot_name} - Offline - Not Scheduled"
                    )
                else:
                    self.update_scheduled_list(f"{robot_name} - Online")

    def is_robot_alive(self, ip_address):

        try:
            response = ping(ip_address, count=4, timeout=0.8)
            return response.success()
        except Exception as e:
            if self.logger:
                self.logger.error(f"Error occurred while pinging {ip_address}: {e}")

            print(f"Error occurred while pinging {ip_address}: {e}")
            return False

    def remove_robot_info(self, robot_name):

        self.robot_info = [
            (name, ip) for name, ip in self.robot_info if name != robot_name
        ]

    def read_robot_info(self, file_path):

        try:
            self.robot_info = []
            with open(file_path, "r") as file:
                for line in file:
                    if line.startswith("["):
                        continue
                    parts = line.strip().split(":")
                    if len(parts) == 2:
                        robot_name, ip_address = parts
                        self.robot_info.append((robot_name, ip_address))
        except Exception as e:
            if self.logger:
                self.logger.error(f"Error reading file: {str(e)}")

    def terminate_threads(self):

        # Call a method in your Worker class to stop all running threads
        self.worker.terminate_all_threads()

    def check_create_folder(self, path):

        try:
            with open(path, "r") as file:
                content = file.read()
                self.Newpath = re.findall(r"\[(.*?)\]", content)

                return self.Newpath
        except Exception as e:
            if self.logger:
                self.logger.error(f"Error reading file: {str(e)}")

    def move_folders_to_archive(self, main_folder_path, archive_folder_path):

        # Get a list of all folders except 'Archive'
        folders_to_move = [
            folder
            for folder in os.listdir(main_folder_path)
            if os.path.isdir(os.path.join(main_folder_path, folder))
            and folder != "Archive"
        ]

        # Find the latest Rev folder
        latest_rev = 0
        for folder in os.listdir(archive_folder_path):
            if folder.startswith("Rev") and folder[3:].isdigit():
                rev_number = int(folder[3:])
                if rev_number > latest_rev:
                    latest_rev = rev_number

        # Increment the latest Rev number
        next_rev = latest_rev + 1

        # Move each folder to the Archive folder
        for folder in folders_to_move:
            folder_path = os.path.join(main_folder_path, folder)
            destination_path = os.path.join(
                archive_folder_path, f"Rev{next_rev}", folder
            )

            # Create the next Rev folder if it doesn't exist
            if not os.path.exists(os.path.join(archive_folder_path, f"Rev{next_rev}")):
                os.makedirs(os.path.join(archive_folder_path, f"Rev{next_rev}"))

            shutil.move(folder_path, destination_path)

    def start_backup(self):

        global TotalRobots
        global ThreadCount
        self.compile_robot_info()
        if not self.robot_info == []:

            self.tab_widget.setCurrentIndex(1)
            ThreadCount = self.update_thread_count()
            self.worker.update_semaphore(ThreadCount)
            self.thread_count_combobox.setEnabled(False)
            self.backup_button.setEnabled(False)
            self.open_folder_button.setEnabled(False)
            try:
                self.backup_button.setEnabled(False)
                # self.file_button.setEnabled(False)
                self.scheduled_list_widget.clear()
                self.completed_list_widget.clear()

                try:
                    self.main_folder = "".join(self.backup_path)
                    self.archive = self.main_folder + "\\" + "Archive"

                    if not os.path.exists(self.main_folder):
                        os.makedirs(self.main_folder, exist_ok=True)

                    else:
                        os.makedirs(self.archive, exist_ok=True)
                        self.move_folders_to_archive(self.main_folder, self.archive)
                except Exception as e:
                    if self.logger:
                        self.logger.error(f"Error creating main folder: {str(e)}")

                selected_extensions = [
                    checkbox.text()
                    for checkbox in self.extension_checkboxes
                    if checkbox.isChecked()
                ]
                # print(f"from gui: {selected_extensions}")
                if (
                    selected_extensions == ["*.* All Files"]
                    or selected_extensions == []
                ):
                    selected_extensions = (
                        "."  # Set selected_extensions to None to indicate all files
                    )
                else:
                    selected_extensions = [
                        checkbox.text()
                        for checkbox in self.extension_checkboxes
                        if checkbox.isChecked()
                    ]
                # print(f"selected extension: {selected_extensions}")
                for robot_name, ip_address in self.robot_info:
                    if ip_address:
                        self.worker.queue_backup(
                            robot_name,
                            ip_address,
                            self.main_folder,
                            selected_extensions,
                        )
                        TotalRobots += 1
                        self.addProgressBarItem(f"{robot_name}--Scheduled", 0)
                        self.worker.start_next_backup()  # Start the next backup

            except AttributeError:
                print("Please select a file first.")
            except Exception as e:
                if self.logger:
                    self.logger.error(f"Error starting backup: {str(e)}")

                print(f"Error: {str(e)}")

        else:
            QMessageBox.warning(self, "Error", "Please select Robots first.")

    @pyqtSlot(str, int, str, str)
    def update_progress(self, ftp_host, progress, file_name, error):

        global TotalRobots
        item_text = f"{ftp_host} - Running: {progress}% - {file_name} {error}"

        # Check if there's an existing running item for the same FTP host
        running_item = None
        for i in range(self.scheduled_list_widget.count()):
            item = self.scheduled_list_widget.item(i)
            widget = self.scheduled_list_widget.itemWidget(item)
            if widget and f"{ftp_host} - Running" in widget.getText():
                running_item = item
                break

        # If a running item is found, update it; otherwise, add a new one
        if running_item:
            # running_item.setText(item_text)
            self.updateProgressBarItem(running_item, f"{ftp_host} - Running ", progress)
            if "Terminated" in item_text:
                self.remove_scheduled_list_failed(ftp_host, error)
                TotalRobots -= 1
        else:
            # Check if there's an existing scheduled item for the same FTP host
            scheduled_item = None
            for i in range(self.scheduled_list_widget.count()):
                item = self.scheduled_list_widget.item(i)
                widget = self.scheduled_list_widget.itemWidget(item)
                if widget.getText().startswith(f"{ftp_host}"):
                    scheduled_item = item
                    break

            # If a scheduled item is found, update it; otherwise, add a new one
            if scheduled_item and not "Terminated" in item_text:
                self.updateProgressBarItem(scheduled_item, f"{ftp_host} - Running", 0)
                # scheduled_item.setText(item_text)
            elif "Terminated" in item_text and not running_item:
                self.remove_scheduled_list_failed(ftp_host, error)
                TotalRobots -= 1
            else:
                self.addProgressBarItem(item_text, progress)
        # print(TotalRobots)
        # Remove completed item from both scheduled and running lists
        if progress == 100 and file_name == "Completed":
            self.remove_scheduled_list(ftp_host)
            TotalRobots -= 1
        # print(TotalRobots)
        if TotalRobots == 0:
            self.open_folder_button.setEnabled(True)
            self.open_folder_button.setVisible(True)
            # self.scheduled_list_widget.addItem(
            #     "Backup - Completed"
            # )
            self.backup_button.setEnabled(True)
            self.thread_count_combobox.setEnabled(True)
            # self.file_button.setEnabled(False)
            try:
                os.startfile(self.main_folder)  # For Windows
            except Exception as e:
                print(f"Error opening folder: {str(e)}")

    def addProgressBarItem(self, text, value):
        item = QListWidgetItem(self.scheduled_list_widget)
        progressBarItem = ProgressBarListItem(text, value)

        item.setSizeHint(progressBarItem.sizeHint())
        self.scheduled_list_widget.addItem(item)
        self.scheduled_list_widget.setItemWidget(item, progressBarItem)

        return item

    def updateProgressBarItem(self, item, text, value):
        progressBarItem = self.scheduled_list_widget.itemWidget(item)
        if progressBarItem:
            progressBarItem.updateProgress(text, value)

    def update_completed_list(self, ftp_host):
        # Create a QWidget as the container
        widget = QWidget()
        layout = QHBoxLayout(widget)

        # Create a QLabel to display the text
        label = QLabel(f"{ftp_host} - Completed")
        label.setStyleSheet(
            "background-color: #7ABA78; color: white;"
            "border-radius: 5px; padding: 2px 5px;"
        )

        # Add the label to the layout
        layout.addWidget(label)
        layout.setContentsMargins(2, 1, 2, 1)  # Remove margins if needed
        widget.setLayout(layout)

        # Create a QListWidgetItem
        item = QListWidgetItem(self.completed_list_widget)

        # Set the custom widget for the item
        self.completed_list_widget.setItemWidget(item, widget)
        item.setSizeHint(
            widget.sizeHint()
        )  # Important to ensure the item has the correct size

    def update_failed_list(self, ftp_host, error):
        # Create a QWidget as the container
        widget = QWidget()
        layout = QHBoxLayout(widget)

        # Create a QLabel to display the text
        label = QLabel(f"{ftp_host} - Failed Due to : {error}")
        label.setStyleSheet(
            "background-color: #F2613F; color: white;"
            "border-radius: 5px; padding: 2px 5px;"
        )

        # Add the label to the layout
        layout.addWidget(label)
        layout.setContentsMargins(2, 1, 2, 1)  # Remove margins if needed
        widget.setLayout(layout)

        # Create a QListWidgetItem
        item = QListWidgetItem(self.completed_list_widget)

        # Set the custom widget for the item
        self.completed_list_widget.setItemWidget(item, widget)
        item.setSizeHint(
            widget.sizeHint()
        )  # Important to ensure the item has the correct size

    def update_scheduled_list(self, ftp_host):

        self.scheduled_list_widget.addItem(f"{ftp_host}")

    def remove_scheduled_list(self, ftp_host):

        # Also remove the corresponding running item
        for index in range(self.scheduled_list_widget.count()):
            item = self.scheduled_list_widget.item(index)
            widget = self.scheduled_list_widget.itemWidget(item)
            if widget and f"{ftp_host}" in widget.getText():
                row = self.scheduled_list_widget.row(item)
                self.scheduled_list_widget.takeItem(row)
                self.update_completed_list(ftp_host)

                break

    def remove_scheduled_list_failed(self, ftp_host, error):

        # Also remove the corresponding scheduled item
        for index in range(self.scheduled_list_widget.count()):
            item = self.scheduled_list_widget.item(index)
            widget = self.scheduled_list_widget.itemWidget(item)
            if widget and f"{ftp_host}" in widget.getText():
                row = self.scheduled_list_widget.row(item)
                self.scheduled_list_widget.takeItem(row)
                self.update_failed_list(ftp_host, error)

                break


class ProgressBarListItem(QWidget):
    def __init__(self, text, value, parent=None):
        super().__init__(parent)
        self.layout = QHBoxLayout()

        self.label = QLabel(text)
        self.progressBar = QProgressBar()
        self.progressBar.setValue(value)
        self.progressBar.setFixedSize(200, 15)

        self.layout.addWidget(self.label)
        self.progressBarLayout = QHBoxLayout()
        self.progressBarLayout.setContentsMargins(
            0, 0, 0, 0
        )  # Set left margin to adjust start position
        self.progressBarLayout.addWidget(self.progressBar)
        self.layout.setContentsMargins(0, 5, 0, 5)

        self.layout.addLayout(self.progressBarLayout)

        self.setLayout(self.layout)

    def updateProgress(self, text, value):
        self.label.setText(text)
        self.progressBar.setValue(value)

    def getText(self):
        return self.label.text()


class ProjectConfigEditor(QDialog):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("Project Editor")
        self.setFixedSize(300, 500)
        self.initUI()

    def initUI(self):

        layout = QVBoxLayout(self)
        self.backup_directory_label = QLabel()
        layout.addWidget(self.backup_directory_label)
        button_layout = QHBoxLayout()

        self.backup_directory_button = QPushButton("Select Backup Folder")
        self.backup_directory_button.setToolTip(
            "Location of the Folder where all the backup will be stored."
        )
        self.backup_directory_button.clicked.connect(self.selectBackupFolder)
        button_layout.addWidget(self.backup_directory_button)

        self.import_excel_button = QPushButton("Import from Excel")
        self.import_excel_button.setToolTip(
            "Import Robot IP address and name from the Enet Matrix Excel file.\n"
            "This function also supports DCDL files."
        )
        self.import_excel_button.clicked.connect(self.importfromExcelDCDL)
        button_layout.addWidget(self.import_excel_button)

        # Add the horizontal layout to the main vertical layout
        layout.addLayout(button_layout)

        self.robots_table = QTableWidget()
        self.robots_table.setColumnCount(2)
        self.robots_table.setHorizontalHeaderLabels(["Robot Name", "IP"])
        layout.addWidget(self.robots_table)

        # Add initial empty row
        self.addEmptyRow()

        button_layout1 = QHBoxLayout()
        self.add_row_button = QPushButton("Add Row")
        self.add_row_button.setToolTip(
            "Add empty row to the table for the manual entry."
        )
        self.add_row_button.clicked.connect(self.addEmptyRow)
        button_layout1.addWidget(self.add_row_button)

        self.save_button = QPushButton("Save")
        self.save_button.setToolTip("Save project to the default directory.")
        self.save_button.clicked.connect(self.saveConfig)
        button_layout1.addWidget(self.save_button)

        layout.addLayout(button_layout1)

    def selectBackupFolder(self):
        folder_path_project = QFileDialog.getExistingDirectory(
            self, "Select Backup Folder", "", QFileDialog.ShowDirsOnly
        )
        if folder_path_project:
            self.backup_directory_label.setText(
                "Backup Directory: " + folder_path_project
            )

    def addEmptyRow(self):
        rowPosition = self.robots_table.rowCount()
        self.robots_table.insertRow(rowPosition)

        name_edit = QLineEdit()
        ip_edit = QLineEdit()

        self.robots_table.setCellWidget(rowPosition, 0, name_edit)
        self.robots_table.setCellWidget(rowPosition, 1, ip_edit)

    def saveConfig(self, project_name=None):
        backup_directory = self.backup_directory_label.text()

        # Extract the selected directory path from the label text
        backup_directory = backup_directory.replace("Backup Directory: ", "")

        # Check if backup directory is empty
        if not backup_directory:
            QMessageBox.warning(self, "Error", "Backup directory cannot be empty.")
            return

        robots = {}
        for row in range(self.robots_table.rowCount()):
            name_item = self.robots_table.cellWidget(row, 0)
            ip_item = self.robots_table.cellWidget(row, 1)
            name = name_item.text()
            ip = ip_item.text()
            name = name.upper()
            # Ignore blank robot names and corresponding blank IP addresses
            if not name.strip() and not ip.strip():
                continue
            if not self.validateName(name):
                QMessageBox.warning(self, "Error", "Invalid name: '{}'.".format(name))
                return
            if not self.validateIP(ip):
                QMessageBox.warning(
                    self, "Error", "Invalid IP address: '{}'.".format(ip)
                )
                return
            # Check for duplicate names and IPs
            if name in robots:
                QMessageBox.warning(self, "Error", "Duplicate name: '{}'.".format(name))
                return
            if ip in robots.values():
                QMessageBox.warning(
                    self, "Error", "Duplicate IP address: '{}'.".format(ip)
                )
                return
            robots[name] = ip

        if not project_name:  # If project name is not provided, ask for it

            folder_path_project = os.path.join(
                "C:\\Users\\" + getpass.getuser() + "\\Documents\\Exodus"
            )

            if not os.path.exists(folder_path_project):
                os.makedirs(folder_path_project)
            project_name, _ = QFileDialog.getSaveFileName(
                self, "Save Project", folder_path_project, "Project Files (*.ini)"
            )
            if not project_name:  # User canceled the dialog
                QMessageBox.warning(
                    self, "Error", "Please provide a valid project name."
                )
                return

        config = configparser.ConfigParser()
        config["General"] = {"BackupDirectory": backup_directory}
        if "Robots" not in config:
            config["Robots"] = {}
        for name, ip in robots.items():
            config["Robots"][name] = ip

        file_path = os.path.join(folder_path_project, project_name)
        with open(file_path, "w") as configfile:
            config.write(configfile)

        self.accept()

    def validateIP(self, ip):
        # Regular expression for validating IP address
        ip_regex = (
            r"^(?:25[0-5]|2[0-4][0-9]|[01]?[0-9][0-9]?)\.(?:25[0-5]|2[0-4][0-9]|[01]?[0-9][0-9]?)\.(?:25["
            r"0-5]|2[0-4][0-9]|[01]?[0-9][0-9]?)\.(?:25[0-5]|2[0-4][0-9]|[01]?[0-9][0-9]?)$"
        )
        return re.match(ip_regex, ip) is not None

    def validateName(self, name):
        # Regular expressions for validating robot names
        name_regex1 = r"^[A-Za-z]{2}[0-9]{3}[Rr]{1}[0-9]{2}[Bb]{1}[0-9]{2}$"
        name_regex2 = (
            r"^[A-Za-z]{2}[0-9]{3}[Pp]{1}[0-9]{1}[Rr]{1}[0-9]{2}[Bb]{1}[0-9]{2}$"
        )
        name_regex3 = r"^[0-9]{3}[Rr]{1}[0-9]{1}$"
        name_regex4 = r"^[0-9]{4}[Rr]{1}[0-9]{1}$"
        return (
            bool(re.match(name_regex1, name))
            or bool(re.match(name_regex2, name))
            or bool(re.match(name_regex3, name))
            or bool(re.match(name_regex4, name))
        )

    def editProject(self):
        # Get the project file to edit
        project_file, _ = QFileDialog.getOpenFileName(
            self, "Open Project", "", "Project Files (*.ini)"
        )
        if not project_file:  # User canceled the dialog
            return

        # Load configuration from the selected project file
        config = configparser.ConfigParser()
        config.read(project_file)

        # Extract backup directory from the loaded configuration
        backup_directory = config.get("General", "BackupDirectory", fallback="")

        # Update UI with loaded backup directory
        self.backup_directory_label.setText("Backup Directory: " + backup_directory)

        # Clear existing rows from the table
        self.robots_table.setRowCount(0)

        # Populate table with robots from the loaded configuration
        for name, ip in config.items("Robots"):
            rowPosition = self.robots_table.rowCount()
            self.robots_table.insertRow(rowPosition)

            name_edit = QLineEdit(name)
            ip_edit = QLineEdit(ip)

            self.robots_table.setCellWidget(rowPosition, 0, name_edit)
            self.robots_table.setCellWidget(rowPosition, 1, ip_edit)

    def importfromExcelDCDL(self):
        excel_file_path, _ = QFileDialog.getOpenFileName(
            self, "Open Excel File", "", "Excel Files (*.xlsm *.xlsx)"
        )
        if excel_file_path:
            # Show progress dialog while opening Excel file
            progress_dialog = QProgressDialog("Processing File...", None, 0, 0, self)
            progress_dialog.setWindowModality(Qt.WindowModal)
            progress_dialog.setWindowTitle("Processing Data")
            progress_dialog.setCancelButton(None)
            progress_dialog.setMinimumDuration(0)
            progress_dialog.show()
            progress_dialog.setValue(10)
            QApplication.processEvents()
            try:
                # Load workbook
                wb = openpyxl.load_workbook(excel_file_path)

                # Check for "Start Here" sheet
                if "Start Here" in wb.sheetnames:

                    sheet_names = wb.sheetnames
                    sheets_with_keyword = []

                    # Find sheets with the keyword
                    for sheet_name in sheet_names:
                        sheet = wb[sheet_name]
                        for row in sheet.iter_rows(
                            min_row=2, min_col=4, max_col=4, values_only=True
                        ):
                            if (
                                "Robot Fanuc Global" in row
                                or "Robot Fanuc Global (W/Vision)" in row
                            ):
                                sheets_with_keyword.append(sheet_name)
                                break

                    if not sheets_with_keyword:
                        QMessageBox.warning(
                            self, "Warning", "No sheets contain the Robot in column D."
                        )
                        return

                    # Close the progress dialog
                    progress_dialog.close()

                    # Show custom dialog to select sheets
                    dialog = SelectSheetsDialog(sheets_with_keyword)
                    if dialog.exec_():
                        selected_sheets = dialog.selectedSheets()
                        for sheet_name in selected_sheets:
                            sheet = wb[sheet_name]
                            self.loadSheetDataDCDL(sheet)
                        self.robots_table.removeRow(0)

                elif "WorkBook_Setup" in wb.sheetnames:
                    pattern1 = r"^[0-9]{3}[Rr]{1}[0-9]{1}$"
                    pattern3 = r"^[0-9]{2}[Rr]{1}[0-9]{1}$"
                    pattern2 = r"^[0-9]{4}[Rr]{1}[0-9]{1}$"
                    pattern = f"{pattern1}|{pattern2}|{pattern3}"
                    enet_sheet = wb["ENET Matrix"]

                    if enet_sheet["D18"].value == "Code":
                        vlan_cells = [
                            enet_sheet["B13"],
                            enet_sheet["H13"],
                            enet_sheet["M13"],
                            enet_sheet["R13"],
                            enet_sheet["W13"],
                        ]
                        columns = {"name": [2, 8, 13, 18, 23], "ip": [4, 9, 14, 19, 24]}
                        pass
                    else:
                        vlan_cells = [
                            enet_sheet["B13"],
                            enet_sheet["G13"],
                            enet_sheet["L13"],
                            enet_sheet["Q13"],
                            enet_sheet["V13"],
                        ]
                        columns = {"name": [2, 7, 12, 17, 22], "ip": [3, 8, 13, 18, 23]}
                        pass
                    # Extract VLAN data from specified cells

                    vlan_data = {}
                    previous_vlan_name = None

                    for idx, vlan_cell in enumerate(vlan_cells):
                        # Ensure vlan_cell.value is not None
                        if vlan_cell.value is not None:
                            vlan_name = vlan_cell.value[:8]
                        else:
                            # Handle the case when vlan_cell.value is None
                            vlan_name = None

                        if vlan_name:
                            vlan_data.setdefault(vlan_name, [])
                            previous_vlan_name = vlan_name
                        else:
                            vlan_name = previous_vlan_name

                        for row in enet_sheet.iter_rows(
                            min_row=20, max_row=enet_sheet.max_row
                        ):
                            name = row[columns["name"][idx]].value
                            ip = row[columns["ip"][idx]].value
                            if name is not None:
                                if re.match(pattern, name):
                                    vlan_data[vlan_name].append((name, ip))

                    if not vlan_data:
                        QMessageBox.warning(
                            self,
                            "Warning",
                            "No valid VLAN data found in the 'Enet matrix' sheet.",
                        )
                        progress_dialog.close()  # Close progress dialog if there's a warning
                        return

                    # Close the progress dialog
                    progress_dialog.close()

                    # Filter out empty values from vlan_data dictionary
                    filtered_vlan_data = {
                        key: value for key, value in vlan_data.items() if value
                    }

                    # Pass the filtered dictionary to the SelectVlanDialog
                    dialog = SelectVlanDialog(filtered_vlan_data)
                    if dialog.exec_():
                        selected_vlan = dialog.selectedVlan()
                        selected_robots = vlan_data[selected_vlan]
                        for name, ip in selected_robots:
                            if self.validateIP(ip):
                                self.addRobotRow(name, ip)
                        self.robots_table.removeRow(0)
                else:
                    QMessageBox.warning(
                        self,
                        "Error",
                        "The Excel file does not contain a Data or file is invalid.",
                    )
                    progress_dialog.close()
                    return

            except Exception as e:
                traceback.print_exc()
                QMessageBox.warning(
                    self, "Error", "Failed to read Excel file: {}".format(e)
                )

    def loadSheetDataDCDL(self, sheet):
        try:
            for row in sheet.iter_rows(min_row=2, values_only=True):
                keyword = row[3]  # Assuming keyword is in column D
                name = row[2]  # Assuming name is in column C
                ip = row[5]  # Assuming IP address is in column F
                if (
                    (
                        keyword == "Robot Fanuc Global (W/Vision)"
                        or keyword == "Robot Fanuc Global"
                    )
                    and name
                    and ip
                ):
                    self.addRobotRow(name, ip)
        except Exception as e:
            QMessageBox.warning(
                self, "Error", "Failed to load data from sheet: {}".format(e)
            )

    def addRobotRow(self, name, ip):
        rowPosition = self.robots_table.rowCount()
        self.robots_table.insertRow(rowPosition)
        name_edit = QLineEdit(name)
        ip_edit = QLineEdit(ip)
        self.robots_table.setCellWidget(rowPosition, 0, name_edit)
        self.robots_table.setCellWidget(rowPosition, 1, ip_edit)


class SelectSheetsDialog(QDialog):
    def __init__(self, sheet_names):
        super().__init__()
        self.setWindowTitle("Select Sheets")
        layout = QVBoxLayout(self)
        self.sheet_list = QListWidget()
        self.sheet_list.setSelectionMode(QAbstractItemView.MultiSelection)
        self.sheet_list.addItems(sheet_names)
        layout.addWidget(self.sheet_list)
        buttons_layout = QHBoxLayout()
        self.ok_button = QPushButton("OK")
        self.ok_button.clicked.connect(self.accept)
        self.cancel_button = QPushButton("Cancel")
        self.cancel_button.clicked.connect(self.reject)
        buttons_layout.addWidget(self.ok_button)
        buttons_layout.addWidget(self.cancel_button)
        layout.addLayout(buttons_layout)

    def selectedSheets(self):
        return [item.text() for item in self.sheet_list.selectedItems()]


class SelectVlanDialog(QDialog):
    def __init__(self, vlan_data):
        super().__init__()
        self.setWindowTitle("Select VLAN")
        layout = QVBoxLayout(self)
        self.vlan_list = QListWidget()
        self.vlan_list.setSelectionMode(QAbstractItemView.SingleSelection)
        for vlan in vlan_data.keys():
            self.vlan_list.addItem(vlan)
        layout.addWidget(self.vlan_list)
        buttons_layout = QHBoxLayout()
        self.ok_button = QPushButton("OK")
        self.ok_button.clicked.connect(self.accept)
        self.cancel_button = QPushButton("Cancel")
        self.cancel_button.clicked.connect(self.reject)
        buttons_layout.addWidget(self.ok_button)
        buttons_layout.addWidget(self.cancel_button)
        layout.addLayout(buttons_layout)

    def selectedVlan(self):
        selected_items = self.vlan_list.selectedItems()
        if selected_items:
            return selected_items[0].text()
        return None


class Worker(QObject):
    progress_signal = pyqtSignal(str, int, str, str)

    def __init__(self):
        super().__init__()
        exe_dir = os.path.dirname(sys.executable)
        log_file_path = os.path.join(exe_dir, "Exodus_log.dmp")
        logging.basicConfig(
            filename=log_file_path,
            level=logging.DEBUG,
            format="%(asctime)s - %(levelname)s - %(message)s",
        )
        self.logger = logging.getLogger(__name__)
        self.logger.info(f"Started From {os.environ.get("COMPUTERNAME", "")}")
        global ThreadCount
        self.mutex = QMutex()
        self.queue = []

    def update_semaphore(self, ThreadCount):
        """
        Updates the semaphore of the worker based on the given ThreadCount.

        :param ThreadCount: An integer representing the number of simultaneous backups allowed.
        :return: None
        """
        self.semaphore = threading.Semaphore(
            ThreadCount
        )  # Default limit to 10 simultaneous backups

    def start_backup(self, robot_name, ftp_host, main_folder, selected_extensions):
        """
        Starts the backup process for a robot.

        Parameters:
            self (object): The instance of the class.
            robot_name (str): The name of the robot to backup.
            ftp_host (str): The FTP host to connect to for backing up.
            main_folder (str): The main folder where backups are stored.
            selected_extensions (list): The list of selected file extensions to backup.

        Returns:
            None
        """
        try:
            thread = threading.Thread(
                target=self.backup_robot,
                args=(robot_name, ftp_host, main_folder, selected_extensions),
            )
            thread.daemon = True
            thread.start()
        except Exception as e:
            if self.logger:
                self.logger.error(f"Error starting backup: {str(e)}")

    def queue_backup(self, robot_name, ftp_host, main_folder, selected_extensions):
        """
        Queues a backup for a robot.

        Args:
            robot_name (str): The name of the robot to backup.
            ftp_host (str): The FTP host to connect to for backing up.
            main_folder (str): The main folder where backups are stored.
            selected_extensions (list): The list of selected file extensions to backup.

        Raises:
            Exception: If there is an error queuing the backup.

        Returns:
            None
        """
        try:
            self.mutex.lock()
            self.queue.append((robot_name, ftp_host, main_folder, selected_extensions))
            self.mutex.unlock()
        except Exception as e:
            if self.logger:
                self.logger.error(f"Error queuing backup: {str(e)}")

    def start_next_backup(self):
        """
        Starts the next backup process.

        This function retrieves the next backup host from the queue and starts the backup process.

        Parameters:
            self (object): The instance of the class.

        Returns:
            None

        Raises:
            Exception: If there is an error starting the next backup.
        """
        try:
            next_host = self.get_next_backup()
            if next_host:
                self.start_backup(*next_host)
        except Exception as e:
            if self.logger:
                self.logger.error(f"Error starting next backup: {str(e)}")

    def get_next_backup(self):
        """
        A description of the function that tries to get the next backup from the queue.

        This function attempts to retrieve the next backup host from the queue.
        It locks the mutex before accessing the queue to ensure thread safety.
        If the queue is not empty, it removes the first item from the queue and returns it as the next host.
        After retrieving the host, it unlocks the mutex.
        If an exception is encountered during this process, the function logs an error using the logger.

        Returns:
            The next host if available, None otherwise.
        """
        try:
            self.mutex.lock()
            if self.queue:
                next_host = self.queue.pop(0)
                self.mutex.unlock()
                return next_host
            self.mutex.unlock()
        except Exception as e:
            if self.logger:
                self.logger.error(f"Error getting next backup: {str(e)}")
        return None

    def terminate_all_threads(self):
        """
        Terminate all active threads.

        This function terminates all active threads in the current object.
        It iterates over each thread in the `active_threads` list, terminates the thread,
        and waits for it to terminate. After terminating all threads,
        the `active_threads` list is cleared.

        Parameters:
            self (object): The current object.

        Returns:
            None
        """
        for thread in self.active_threads:
            thread.terminate()  # Terminate the thread
            thread.join()  # Wait for the thread to terminate
        self.active_threads.clear()  # Clear the list of active threads

    def backup_robot(self, robot_name, ftp_host, main_folder, selected_extensions):
        """
        Backup robot files from an FTP server.

        Args:
            robot_name (str): The name of the robot to backup.
            ftp_host (str): The FTP host to connect to for backing up.
            main_folder (str): The main folder where backups are stored.
            selected_extensions (list): The list of selected file extensions to backup.

        Returns:
            None
        """
        try:
            self.semaphore.acquire()  # Acquire semaphore slot
            now = datetime.now().strftime("%Y-%m-%d_%HH_%MM")
            robot_folder = main_folder + "\\" + robot_name + "_" + now
            robot_MD = robot_folder + "\\md"
            with FTP(ftp_host, timeout=10) as ftp:
                ftp.login()  # Add your credentials if needed
                ftp.set_pasv(True)
                # Create individual folder for robot within the main folder
                # Create main folder with current date and time on desktop
                os.makedirs(robot_folder, exist_ok=True)
                # Download files from the robot to its folder
                total_files = self.count_files(ftp, "/md:", selected_extensions)
                self.progress_signal.emit(
                    robot_name, 0, "", ""
                )  # Signal to update scheduled list
                os.makedirs(robot_MD, exist_ok=True)
                self.download_directory(
                    ftp,
                    "/md:",
                    str(robot_MD),
                    total_files,
                    robot_name,
                    selected_extensions,
                )  # Adjust path as needed
                self.progress_signal.emit(robot_name, 100, "Completed", "")
        except Exception as e:
            if os.path.exists(robot_folder):
                shutil.rmtree(robot_folder)  # Delete the main folder for the robot
            self.progress_signal.emit(robot_name, 0, "Terminated", str(e))
            if self.logger:
                self.logger.error(f"Error backing up {robot_name}: {str(e)}")
                # Iterate through active threads and terminate the one related to the current robot
                for thread in threading.enumerate():
                    if thread.name == robot_name:
                        thread.terminate()  # Terminate the thread for the specific robot
                        self.progress_signal.emit(robot_name, 0, "Terminated", str(e))
                        break
            print(f"Error backing up {robot_name}: {str(e)}")
        finally:
            self.semaphore.release()  # Release semaphore slot

    def count_files(self, ftp, path, selected_extensions):
        """
        Counts the number of files in the given FTP directory that have selected extensions.

        :param ftp: An instance of FTP class for the FTP connection.
        :type ftp: ftplib.FTP
        :param path: The path of the directory to count files in.
        :type path: str
        :param selected_extensions: A list of file extensions to count.
        :type selected_extensions: List[str]
        :return: The number of files in the directory with selected extensions.
        :rtype: int
        """
        ftp.cwd(path)
        items = ftp.nlst()
        file_count = 0
        for item in items:
            if self.is_file(ftp, item, selected_extensions) and any(
                ext in item for ext in selected_extensions
            ):
                file_count += 1
        ftp.cwd("..")
        return file_count

    def download_directory(
        self, ftp, path, local_path, total_files, robot_name, selected_extensions
    ):
        """
        Downloads a directory from an FTP server and saves it locally.

        Args:
            ftp (FTP): The FTP connection object.
            path (str): The remote directory path to download.
            local_path (str): The local directory path to save the downloaded files.
            total_files (int): The total number of files to download.
            robot_name (str): The name of the robot performing the download.
            selected_extensions (List[str]): The list of file extensions to download.

        Returns:
            None
        """
        ftp.cwd(path)
        items = ftp.nlst()
        completed_files = 0
        for item in items:
            local_file_path = os.path.join(local_path, item)
            if self.is_file(ftp, item, selected_extensions):
                with open(local_file_path, "wb") as f:
                    ftp.retrbinary(f"RETR {item}", f.write, blocksize=64 * 1024)
                completed_files += 1
                progress = int((completed_files / total_files) * 100)
                self.progress_signal.emit(
                    robot_name, progress, item, ""
                )  # Emit progress signal

            else:
                # Skip creating a directory for files with extensions
                if "." not in item:
                    subdirectory = os.path.join(local_path, item)
                    os.makedirs(subdirectory, exist_ok=True)
                    self.download_directory(
                        ftp,
                        item,
                        subdirectory,
                        total_files,
                        robot_name,
                        selected_extensions,
                    )
        ftp.cwd("..")

    def is_file(self, ftp, item, selected_extensions):
        """
        A function to determine if the given item is a file based on selected extensions.

        :param ftp: The FTP connection object.
        :param item: The item to check if it's a file.
        :param selected_extensions: The list of file extensions to consider.

        :return: True if the item is a file based on the selected extensions, False otherwise.
        """
        if "." in item:
            if selected_extensions == ".":
                return True
            else:
                extension = item.split(".")[1]
                extension = "." + extension
                if extension in selected_extensions:
                    return True
                else:
                    return False
        else:
            try:
                # Only navigate to directory if it doesn't have an extension
                ftp.cwd(item)
                ftp.cwd("..")
                return False
            except:
                return True


if __name__ == "__main__":
    username = getpass.getuser()
    app = QApplication(sys.argv)

    import os

    if not any(
        substring in os.environ.get("COMPUTERNAME", "") for substring in ["MSVN"]
    ):

        QMessageBox.critical(None, "Access Denied", "System not valid")
        sys.exit()
    ## Create a QPixmap with your splash screen image
    splash_pix = QPixmap(resource_path("RobotSplash.png"))
    splash = QSplashScreen(splash_pix, Qt.WindowStaysOnTopHint)
    splash.setMask(splash_pix.mask())
    splash.show()
    app.processEvents()  # Allows UI to update while loading data
    QTimer.singleShot(800, splash.close)
    icon_path = resource_path("MAGNAlogo.png")
    # Set the window icon
    app.setWindowIcon(QIcon(icon_path))
    ex = FTPBackup()
    ex.show()
    sys.exit(app.exec_())
